"""Excel processing utilities for flight school data normalization and export."""

from __future__ import annotations

import json
from collections.abc import Iterable, Sequence
from collections import Counter
from dataclasses import asdict, dataclass, is_dataclass
from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl import load_workbook  # type: ignore[import]
from openpyxl.formatting.rule import CellIsRule  # type: ignore[import]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import]
from openpyxl.utils import get_column_letter  # type: ignore[import]
from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore[import]

from watercrawl.domain.models import (
    EnrichmentResult,
    SchoolRecord,
    normalize_province,
    normalize_status,
)

from .column_inference import ColumnInferenceEngine, ColumnInferenceResult

from . import config  # type: ignore
from .normalization import ColumnNormalizationRegistry, normalize_numeric_value

EXPECTED_COLUMNS = list(config.EXPECTED_COLUMNS)

_SUPPORTED_SUFFIXES = {".csv", ".xlsx", ".xls"}
EVIDENCE_SHEET = "Evidence"


@dataclass(frozen=True)
class WorkbookTheme:
    accent: str
    status_colours: dict[str, str]


def _confidence_to_colour(confidence: int) -> str:
    """Return an ARGB hex colour based on confidence score."""

    if confidence >= 80:
        return "FF93C47D"  # green accent
    if confidence >= 60:
        return "FFF6B26B"  # amber
    if confidence >= 40:
        return "FFF4CCCC"  # light red
    return "FFD9D2E9"  # muted violet fallback


def _resolve_theme() -> WorkbookTheme:
    """Derive workbook theme colours from the active refinement profile."""

    default_confidence = dict(getattr(config, "DEFAULT_CONFIDENCE_BY_STATUS", {}))
    status_colours = {
        status: _confidence_to_colour(score)
        for status, score in default_confidence.items()
    }
    if default_confidence:
        top_status = max(
            default_confidence,
            key=lambda status: default_confidence[status],
        )
        accent = status_colours[top_status]
    else:
        accent = "FF93C47D"
    return WorkbookTheme(accent=accent, status_colours=status_colours)


def _apply_table(worksheet, name: str, *, style: str = "TableStyleMedium9") -> None:
    if worksheet.max_row < 2 or worksheet.max_column < 1:
        return
    last_column = get_column_letter(worksheet.max_column)
    table = Table(
        displayName=name,
        ref=f"A1:{last_column}{worksheet.max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _auto_fit_columns(worksheet) -> None:
    widths: dict[int, int] = {}
    for row in worksheet.iter_rows(values_only=True):
        for index, value in enumerate(row, start=1):
            if value is None:
                continue
            text = str(value)
            widths[index] = max(widths.get(index, 0), len(text) + 2)
    for index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(index)].width = min(width, 60)


def _apply_status_formatting(worksheet, theme: WorkbookTheme) -> None:
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
    status_column = None
    for column_index, cell in enumerate(header_row, start=1):
        if isinstance(cell.value, str) and cell.value.strip().lower() == "status":
            status_column = column_index
            break
    if status_column is None:
        return
    column_letter = get_column_letter(status_column)
    start_row = 2
    end_row = worksheet.max_row
    if end_row < start_row:
        return
    cell_range = f"{column_letter}{start_row}:{column_letter}{end_row}"
    for status, colour in theme.status_colours.items():
        fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{status}"'],
                fill=fill,
            ),
        )


def _format_issues_sheet(worksheet, theme: WorkbookTheme) -> None:
    _apply_table(worksheet, "IssuesTable", style="TableStyleMedium6")
    _auto_fit_columns(worksheet)
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
    severity_column = None
    for column_index, cell in enumerate(header_row, start=1):
        if isinstance(cell.value, str) and cell.value.strip().lower() == "severity":
            severity_column = column_index
            break
    if severity_column is None:
        return
    column_letter = get_column_letter(severity_column)
    start_row = 2
    end_row = worksheet.max_row
    if end_row < start_row:
        return
    warn_fill = PatternFill(start_color="FFF6B26B", end_color="FFF6B26B", fill_type="solid")
    block_fill = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")
    cell_range = f"{column_letter}{start_row}:{column_letter}{end_row}"
    worksheet.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"warn"'], fill=warn_fill),
    )
    worksheet.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"block"'], fill=block_fill),
    )


def _apply_evidence_hyperlinks(worksheet, data_rows: Mapping[int, int]) -> None:
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rowid_index = None
    sources_index = None
    for idx, header in enumerate(headers, start=1):
        if isinstance(header, str) and header.lower() == "rowid":
            rowid_index = idx
        if isinstance(header, str) and header.lower() == "sources":
            sources_index = idx
    if rowid_index is None:
        return
    for row in worksheet.iter_rows(min_row=2):
        row_id_cell = row[rowid_index - 1]
        if row_id_cell.value is None:
            continue
        try:
            row_identifier = int(str(row_id_cell.value))
        except ValueError:
            continue
        target_row = data_rows.get(row_identifier, row_identifier + 2)
        row_id_cell.hyperlink = f"#{config.CLEANED_SHEET}!A{target_row}"  # type: ignore[assignment]
        row_id_cell.style = "Hyperlink"
        if sources_index is not None:
            sources_cell = row[sources_index - 1]
            if sources_cell.value:
                urls = [
                    part.strip()
                    for part in str(sources_cell.value).split(";")
                    if part.strip()
                ]
                if urls:
                    sources_cell.value = "\n".join(urls)
                    sources_cell.hyperlink = urls[0]  # type: ignore[assignment]  # type: ignore[assignment]
                    sources_cell.style = "Hyperlink"
                    sources_cell.alignment = Alignment(wrap_text=True, vertical="top")


def _format_evidence_sheet(worksheet, data_rows: Mapping[int, int]) -> None:
    _apply_table(worksheet, "EvidenceTable", style="TableStyleMedium4")
    _apply_evidence_hyperlinks(worksheet, data_rows)
    _auto_fit_columns(worksheet)
    worksheet.freeze_panes = "A2"


def _build_row_lookup(worksheet) -> dict[int, int]:
    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rowid_index = None
    for idx, value in enumerate(header, start=1):
        if isinstance(value, str) and value.strip().lower() == "rowid":
            rowid_index = idx
            break
    mapping: dict[int, int] = {}
    if rowid_index is not None:
        for excel_row, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            cell_value = row[rowid_index - 1].value
            if cell_value is None:
                continue
            try:
                mapping[int(str(cell_value))] = excel_row
            except ValueError:
                continue
    else:
        for excel_row in range(2, worksheet.max_row + 1):
            mapping[excel_row - 2] = excel_row
    return mapping


def _populate_summary_sheet(
    workbook,
    dataframe: pd.DataFrame,
    issues_count: int,
    evidence_count: int,
    theme: WorkbookTheme,
) -> None:
    sheet = workbook.create_sheet("Summary", 0)
    sheet["A1"] = getattr(config.PROFILE, "name", "Enrichment Summary")
    sheet["A1"].font = Font(size=14, bold=True)
    sheet["A1"].fill = PatternFill(start_color=theme.accent, end_color=theme.accent, fill_type="solid")
    metrics = [
        ("Total Rows", len(dataframe)),
        ("Evidence Entries", evidence_count),
        ("Issues Logged", issues_count),
    ]
    sheet.append(["Metric", "Value"])
    for label, value in metrics:
        sheet.append([label, value])
    metrics_table_end = sheet.max_row
    sheet_tables = Table(
        displayName="SummaryMetrics",
        ref=f"A2:B{metrics_table_end}",
    )
    sheet_tables.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9",
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(sheet_tables)

    status_counts: Counter[str] = Counter()
    if "Status" in dataframe.columns:
        status_counts.update(dataframe["Status"].dropna().astype(str))
    sheet["D2"] = "Status"
    sheet["E2"] = "Count"
    for index, status in enumerate(config.CANONICAL_STATUSES, start=3):
        sheet.cell(row=index, column=4, value=status)
        sheet.cell(row=index, column=5, value=status_counts.get(status, 0))
    status_table_end = 2 + len(config.CANONICAL_STATUSES)
    status_table = Table(
        displayName="SummaryStatuses",
        ref=f"D2:E{status_table_end}",
    )
    status_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight11",
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(status_table)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 10


def _populate_lists_sheet(workbook) -> None:
    sheet = workbook.create_sheet(config.LISTS_SHEET)
    sheet["A1"] = "Statuses"
    for index, status in enumerate(config.CANONICAL_STATUSES, start=2):
        sheet.cell(row=index, column=1, value=status)
    sheet["B1"] = "Provinces"
    for index, province in enumerate(config.PROVINCES, start=2):
        sheet.cell(row=index, column=2, value=province)
    sheet.sheet_state = "hidden"


def _normalize_records(records: Iterable[Any]) -> list[dict[str, Any]]:
    normalised: list[dict[str, Any]] = []
    for record in records:
        if isinstance(record, Mapping):
            normalised.append(dict(record))
            continue
        if hasattr(record, "as_dict"):
            normalised.append(dict(record.as_dict()))
            continue
        if is_dataclass(record) and not isinstance(record, type):
            normalised.append(asdict(record))
            continue
        try:
            normalised.append(dict(record))
        except TypeError:
            normalised.append({"value": record})
    return normalised


def _column_key(name: str) -> str:
    return "".join(ch for ch in name.casefold() if ch.isalnum())


def _expand_input(target: Path) -> list[Path]:
    if target.is_dir():
        return [
            child
            for child in sorted(target.iterdir(), key=lambda item: item.name.lower())
            if child.is_file() and child.suffix.lower() in _SUPPORTED_SUFFIXES
        ]
    return [target]


def _collect_inputs(path_or_paths: Path | Sequence[Path]) -> list[Path]:
    inputs: list[Path] = []
    if isinstance(path_or_paths, Path):
        inputs.extend(_expand_input(path_or_paths))
    else:
        for entry in path_or_paths:
            inputs.extend(_expand_input(entry))
    return inputs


def _resolve_sheet_name(
    path: Path, sheet_map: Mapping[str, str] | None
) -> str:
    if not sheet_map:
        return config.CLEANED_SHEET
    for key in (str(path), path.name, path.stem):
        if key in sheet_map:
            return sheet_map[key]
    return config.CLEANED_SHEET


def _align_columns(
    frame: pd.DataFrame, descriptors: Sequence[Any]
) -> tuple[pd.DataFrame, set[str], ColumnInferenceResult | None]:
    # Late import to avoid circular dependency at module import time.
    from watercrawl.core.profiles import ColumnDescriptor

    if not descriptors:
        return frame, set(), None
    required_columns = set(getattr(config, "EXPECTED_COLUMNS", ()))
    canonical_order: list[str] = []
    alias_lookup: dict[str, str] = {}
    column_descriptors: list[ColumnDescriptor] = []
    for descriptor in descriptors:
        if isinstance(descriptor, ColumnDescriptor):
            canonical_order.append(descriptor.name)
            column_descriptors.append(descriptor)
            alias_lookup[_column_key(descriptor.name)] = descriptor.name
            if descriptor.required:
                required_columns.add(descriptor.name)
            for label in descriptor.candidate_labels():
                alias_lookup[_column_key(label)] = descriptor.name
        else:
            name = getattr(descriptor, "name", None)
            if name:
                canonical = str(name)
                canonical_order.append(canonical)
                alias_lookup[_column_key(canonical)] = canonical

    inference_result: ColumnInferenceResult | None = None
    rename_map: dict[str, str] = {}
    if column_descriptors:
        engine = ColumnInferenceEngine(column_descriptors)
        inference_result = engine.infer(frame)
        rename_map.update(inference_result.rename_map)

    for column in frame.columns:
        source = str(column)
        if source in rename_map:
            continue
        candidate = alias_lookup.get(_column_key(source))
        if candidate and candidate != source and candidate not in rename_map.values():
            rename_map[source] = candidate

    aligned = frame.rename(columns=rename_map).copy()
    missing_columns: set[str] = set()
    for name in canonical_order:
        if name not in aligned.columns:
            aligned[name] = pd.NA
            if name in required_columns:
                missing_columns.add(name)

    ordered_columns = [name for name in canonical_order if name in aligned.columns]
    remaining = [col for col in aligned.columns if col not in ordered_columns]
    return aligned[ordered_columns + remaining], missing_columns, inference_result


class ExcelExporter:
    """Exports enriched dataframes to Excel and CSV artefacts."""

    def __init__(self, workbook_path: Path, provenance_path: Path):
        """Initialize the exporter with paths."""
        self.workbook_path = workbook_path
        self.provenance_path = provenance_path

    def write(
        self,
        enriched_df: pd.DataFrame,
        provenance_rows: Iterable[dict],
        *,
        issues: Iterable[Any] | None = None,
    ) -> None:
        """Write the enriched dataframe, summary sheets, and provenance artefacts."""
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        self.provenance_path.parent.mkdir(parents=True, exist_ok=True)

        evidence_records = _normalize_records(list(provenance_rows))
        issue_source = (
            issues
            if issues is not None
            else enriched_df.attrs.get("quality_issues", [])
        )
        issue_records = _normalize_records(list(issue_source or []))

        issues_frame = (
            pd.DataFrame(issue_records)
            if issue_records
            else pd.DataFrame(
                columns=[
                    "row_id",
                    "organisation",
                    "code",
                    "severity",
                    "message",
                    "remediation",
                ]
            )
        )
        issues_frame = issues_frame.rename(
            columns={
                "row_id": "Row ID",
                "organisation": "Organisation",
                "code": "Code",
                "severity": "Severity",
                "message": "Message",
                "remediation": "Remediation",
            }
        )
        standard_issue_columns = [
            "Row ID",
            "Organisation",
            "Code",
            "Severity",
            "Message",
            "Remediation",
        ]
        issue_remainder = [
            column
            for column in issues_frame.columns
            if column not in standard_issue_columns
        ]
        issues_frame = issues_frame.reindex(
            columns=standard_issue_columns + issue_remainder,
        )
        evidence_frame = (
            pd.DataFrame(evidence_records)
            if evidence_records
            else pd.DataFrame(
                columns=[
                    "RowID",
                    "Organisation",
                    "What changed",
                    "Sources",
                    "Notes",
                    "Timestamp",
                    "Confidence",
                ]
            )
        )
        evidence_columns = [
            "RowID",
            "Organisation",
            "What changed",
            "Sources",
            "Notes",
            "Timestamp",
            "Confidence",
        ]
        evidence_remainder = [
            column
            for column in evidence_frame.columns
            if column not in evidence_columns
        ]
        evidence_frame = evidence_frame.reindex(
            columns=evidence_columns + evidence_remainder,
        )

        with pd.ExcelWriter(self.workbook_path, engine="openpyxl") as writer:
            enriched_df.to_excel(writer, sheet_name=config.CLEANED_SHEET, index=False)
            issues_frame.to_excel(writer, sheet_name=config.ISSUES_SHEET, index=False)
            evidence_frame.to_excel(writer, sheet_name=EVIDENCE_SHEET, index=False)

        evidence_frame.to_csv(self.provenance_path, index=False)

        workbook = load_workbook(self.workbook_path)
        theme = _resolve_theme()

        data_sheet = workbook[config.CLEANED_SHEET]
        data_sheet.freeze_panes = "A2"
        _apply_table(data_sheet, "DataTable", style="TableStyleMedium9")
        _apply_status_formatting(data_sheet, theme)
        _auto_fit_columns(data_sheet)

        issues_sheet = workbook[config.ISSUES_SHEET]
        _format_issues_sheet(issues_sheet, theme)

        evidence_sheet = workbook[EVIDENCE_SHEET]
        row_lookup = _build_row_lookup(data_sheet)
        _format_evidence_sheet(evidence_sheet, row_lookup)

        _populate_summary_sheet(
            workbook,
            enriched_df,
            len(issue_records),
            len(evidence_records),
            theme,
        )
        if config.LISTS_SHEET not in workbook.sheetnames:
            _populate_lists_sheet(workbook)

        workbook.save(self.workbook_path)


def read_dataset(
    path: Path | Sequence[Path],
    *,
    registry: ColumnNormalizationRegistry | None = None,
    sheet_map: Mapping[str, str] | None = None,
) -> pd.DataFrame:
    """Read and normalize a dataset from one or more paths."""

    input_paths = _collect_inputs(path)
    if not input_paths:
        raise ValueError("No supported dataset files were provided")

    descriptors = getattr(config, "COLUMN_DESCRIPTORS", ())
    frames: list[pd.DataFrame] = []
    source_rows: list[dict[str, Any]] = []
    missing_columns_global: set[str] = set()
    inference_results: list[ColumnInferenceResult] = []
    for input_path in input_paths:
        suffix = input_path.suffix.lower()
        sheet_name: str | None = None
        if suffix in {".xlsx", ".xls"}:
            sheet_name = _resolve_sheet_name(input_path, sheet_map)
            frame = pd.read_excel(input_path, sheet_name=sheet_name)
        elif suffix == ".csv":
            frame = pd.read_csv(input_path)
        else:
            raise ValueError(f"Unsupported file format: {suffix}")
        aligned, missing_columns, inference_result = _align_columns(frame, descriptors)
        frames.append(aligned)
        if missing_columns:
            missing_columns_global.update(missing_columns)
        if inference_result is not None:
            inference_results.append(inference_result)
        for local_index, row_index in enumerate(aligned.index):
            source_rows.append(
                {
                    "row": len(source_rows),
                    "path": str(input_path.resolve()),
                    "sheet": sheet_name,
                    "source_row": (
                        int(row_index)
                        if isinstance(row_index, (int, float)) and not pd.isna(row_index)
                        else str(row_index) if row_index is not None else None
                    ),
                    "local_index": local_index,
                }
            )

    combined = pd.concat(frames, ignore_index=True, sort=False)
    metadata_attrs: dict[str, Any] = {
        "source_rows": source_rows,
        "source_files": sorted({entry["path"] for entry in source_rows}),
    }
    if missing_columns_global:
        metadata_attrs["missing_columns"] = sorted(missing_columns_global)
    if sheet_map:
        metadata_attrs["sheet_overrides"] = dict(sheet_map)
    if inference_results:
        summary = ColumnInferenceResult.merge(inference_results)
        metadata_attrs["column_inference"] = summary.to_dict()

    active_registry = registry or getattr(config, "COLUMN_NORMALIZATION_REGISTRY", None)
    diagnostics: dict[str, Any] = {}
    normalized_columns: set[str] = set()

    working_frame = combined
    if active_registry and descriptors:
        working = combined.copy()
        for descriptor in descriptors:
            if descriptor.name not in working.columns:
                continue
            result = active_registry.normalize_series(
                descriptor, working[descriptor.name]
            )
            working[descriptor.name] = result.series
            diagnostics[descriptor.name] = result.diagnostics.to_dict()
            normalized_columns.add(descriptor.name)
        working_frame = working

    remaining_rules: dict[str, dict[str, Any]] | None = None
    if active_registry is not None:
        remaining_rules = {
            name: rule
            for name, rule in active_registry.numeric_rules.items()
            if name not in normalized_columns
        }

    normalized = normalize_numeric_units(working_frame, rules=remaining_rules)
    normalized = normalize_categorical_values(
        normalized, skip_columns=normalized_columns
    )
    normalized.attrs.update(metadata_attrs)

    if diagnostics:
        report_path = config.INTERIM_DIR / "normalization_report.json"
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(json.dumps(diagnostics, indent=2, sort_keys=True))

    return normalized


def write_dataset(df: pd.DataFrame, path: Path) -> None:
    """Write the dataframe to the given path."""
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=config.CLEANED_SHEET, index=False)
        return
    if suffix == ".csv":
        df.to_csv(path, index=False)
        return
    raise ValueError(f"Unsupported file format: {suffix}")


def load_school_records(path: Path = config.SOURCE_XLSX) -> list[SchoolRecord]:
    """Load school records from the dataset at the given path."""
    df = read_dataset(path)
    missing = [col for col in EXPECTED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"Missing expected columns: {missing}")
    return [SchoolRecord.from_dataframe_row(row) for _, row in df.iterrows()]


def append_enrichment_columns(
    df: pd.DataFrame, results: list[EnrichmentResult]
) -> pd.DataFrame:
    """Append enrichment columns to the dataframe."""
    enrichment_rows = [result.record.as_dict() for result in results]
    enrichment_df = pd.DataFrame(enrichment_rows)
    combined = pd.concat([df.reset_index(drop=True), enrichment_df], axis=1)
    return combined


def normalize_numeric_units(
    frame: pd.DataFrame,
    *,
    rules: dict[str, dict[str, Any]] | None = None,
) -> pd.DataFrame:
    """Normalize numeric units in the dataframe according to predefined rules."""

    normalized = frame.copy()
    normalized.attrs.update(frame.attrs)
    rule_lookup = rules if rules is not None else dict(config.NUMERIC_UNIT_LOOKUP)
    if not rule_lookup:
        return normalized

    for column, rule in rule_lookup.items():
        if column not in normalized.columns:
            continue
        allowed_units = set(rule.get("allowed_units", set()))

        def _apply(value: Any) -> Any:
            return normalize_numeric_value(
                value=value,
                column=column,
                rule=rule,
                allowed_units=allowed_units,
            )

        normalized[column] = normalized[column].apply(_apply)

    return normalized


def normalize_categorical_values(
    frame: pd.DataFrame, *, skip_columns: Iterable[str] | None = None
) -> pd.DataFrame:
    """Normalize categorical values in the dataframe."""
    normalized = frame.copy()
    normalized.attrs.update(frame.attrs)
    skip = set(skip_columns or ())
    if "Province" in normalized.columns and "Province" not in skip:
        normalized["Province"] = normalized["Province"].apply(normalize_province)
    if "Status" in normalized.columns and "Status" not in skip:
        normalized["Status"] = normalized["Status"].apply(normalize_status)
    return normalized
