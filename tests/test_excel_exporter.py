from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook  # type: ignore[import]

from watercrawl.core import config, excel


def _build_sample_dataframe() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "RowID": 0,
                "Name of Organisation": "Alpha Flight",
                "Province": "Gauteng",
                "Status": "Verified",
                "Website URL": "https://alpha.example",
                "Contact Person": "Aviator One",
                "Contact Number": "+27123456789",
                "Contact Email Address": "alpha@example.com",
            },
            {
                "RowID": 1,
                "Name of Organisation": "Beta Aviation",
                "Province": "Western Cape",
                "Status": "Needs Review",
                "Website URL": "https://beta.example",
                "Contact Person": "Aviator Two",
                "Contact Number": "+27111222333",
                "Contact Email Address": "beta@example.com",
            },
        ]
    )


def _sample_evidence() -> list[dict[str, str]]:
    return [
        {
            "RowID": "0",
            "Organisation": "Alpha Flight",
            "What changed": "Status updated to Verified",
            "Sources": "https://alpha.example/report; https://alpha.example/news",
            "Notes": "Confirmed via regulator release",
            "Timestamp": "2025-01-01T00:00:00",
            "Confidence": "95",
        },
        {
            "RowID": "1",
            "Organisation": "Beta Aviation",
            "What changed": "Flagged for review",
            "Sources": "https://beta.example/report",
            "Notes": "Awaiting confirmation",
            "Timestamp": "2025-01-02T00:00:00",
            "Confidence": "40",
        },
    ]


def _sample_issues() -> list[dict[str, object]]:
    return [
        {
            "row_id": 1,
            "organisation": "Beta Aviation",
            "code": "missing_phone",
            "severity": "warn",
            "message": "Phone number requires verification",
            "remediation": "Call listed contact",
        }
    ]


def test_excel_exporter_applies_formatting_and_summary(tmp_path: Path) -> None:
    exporter = excel.ExcelExporter(
        tmp_path / "exports" / "enriched.xlsx",
        tmp_path / "exports" / "provenance.csv",
    )
    dataframe = _build_sample_dataframe()

    exporter.write(dataframe, _sample_evidence(), issues=_sample_issues())

    workbook_path = tmp_path / "exports" / "enriched.xlsx"
    workbook = load_workbook(workbook_path)

    assert workbook.sheetnames[0] == "Summary"
    assert config.CLEANED_SHEET in workbook.sheetnames
    assert config.ISSUES_SHEET in workbook.sheetnames
    assert excel.EVIDENCE_SHEET in workbook.sheetnames
    assert config.LISTS_SHEET in workbook.sheetnames

    data_sheet = workbook[config.CLEANED_SHEET]
    assert data_sheet.freeze_panes == "A2"
    assert "DataTable" in data_sheet.tables
    assert len(data_sheet.conditional_formatting) > 0

    summary_sheet = workbook["Summary"]
    assert summary_sheet["A3"].value == "Total Rows"
    assert summary_sheet["B3"].value == len(dataframe)
    assert summary_sheet["D3"].value == config.CANONICAL_STATUSES[0]
    theme = excel._resolve_theme()
    assert summary_sheet["A1"].fill.start_color.rgb == theme.accent

    lists_sheet = workbook[config.LISTS_SHEET]
    assert lists_sheet.sheet_state == "hidden"

    evidence_csv = tmp_path / "exports" / "provenance.csv"
    csv_frame = pd.read_csv(evidence_csv)
    assert not csv_frame.empty
    assert set(csv_frame.columns) >= {"RowID", "Sources", "Confidence"}


def test_excel_exporter_links_evidence_rows(tmp_path: Path) -> None:
    exporter = excel.ExcelExporter(
        tmp_path / "exports" / "enriched.xlsx",
        tmp_path / "exports" / "provenance.csv",
    )
    dataframe = _build_sample_dataframe()

    exporter.write(dataframe, _sample_evidence(), issues=_sample_issues())

    workbook = load_workbook(tmp_path / "exports" / "enriched.xlsx")

    evidence_sheet = workbook[excel.EVIDENCE_SHEET]
    row_cell = evidence_sheet["A2"]
    assert row_cell.hyperlink is not None
    assert row_cell.hyperlink.target == f"#{config.CLEANED_SHEET}!A2"

    sources_cell = evidence_sheet["D2"]
    assert sources_cell.hyperlink is not None
    assert "\n" in sources_cell.value
    assert sources_cell.hyperlink.target.startswith("https://alpha.example")

    issues_sheet = workbook[config.ISSUES_SHEET]
    assert "IssuesTable" in issues_sheet.tables
